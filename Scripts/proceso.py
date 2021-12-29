from autenticarse import loguearse
from altaDocumentos import altaDocs
import openpyxl
from datetime import datetime
from escanearPDFs import recorrer_carpetas
import time

excel = openpyxl.load_workbook("C:/Users/aalarcon/Desktop/osocna.xlsx")
hoja = excel["osocna"]

cantidad = len(hoja["A"])
lineas = []
facturas = []
letra = []
puestoVenta = []
caea = []
vtocaea = []
producto = []
iva = []
precio = []
fecha = []
unidades = []
lineasDup = []
lineasNorm = []

for i in range(2, cantidad + 1):
    letra.append(hoja.cell(row=i, column=4).value)
    facturas.append(hoja.cell(row=i, column=5).value) 
    puestoVenta.append(hoja.cell(row=i, column=3).value) 
    caea.append(hoja.cell(row=i, column=6).value) 
    vtocaea.append(hoja.cell(row=i, column=7).value)
    producto.append(hoja.cell(row=i, column=10).value)
    iva.append(hoja.cell(row=i, column=11).value)
    precio.append(hoja.cell(row=i, column=15).value)
    fecha.append(hoja.cell(row=i, column=2).value) 
    unidades.append(hoja.cell(row=i, column=9).value)

n = 0
c = 0

for f in facturas:
    fecha_convertida = fecha[n]

    descripcion_medicamento = ""
    id_afip = None
    idalicuota = None
    alicuota = None

    if puestoVenta[n] == 44 and producto[n] != None:
        descripcion_medicamento = "Medicamentos"
        id_afip = 1
    elif puestoVenta[n] == 44 and producto[n] == None:
        descripcion_medicamento = "Servicios"
        id_afip = 1
    elif puestoVenta[n] == 50 and producto[n] == None:
        descripcion_medicamento = "Nota de credito servicios"
        id_afip = 3
    elif puestoVenta[n] == 50 and producto[n] != None:
        descripcion_medicamento = "Nota de credito medicamentos"
        id_afip = 3
    
    if iva[n] == 21:
        idalicuota = 5
    else:
        idalicuota = 3

    factRepetida = facturas.count(f)
    if factRepetida > 1:
        c = c + 1
        json = {
            "idAlicuota": idalicuota,
            "cantidad": unidades[n],
            "importeNeto": precio[n],
            "descripcion": producto[n]
        }
        lineasDup.append(json)
        if c == factRepetida:
            altaDocs(letra[n], puestoVenta[n], facturas[n], fecha_convertida, caea[n], descripcion_medicamento, lineasDup, id_afip, idalicuota, iva[n])
            recorrer_carpetas(facturas[n], id_afip, puestoVenta[n], letra[n])
            print("")
            # print(lineasDup)
            c = 0
            lineasDup.clear()
    
    else:
        json2 = {
            "idAlicuota": idalicuota,
            "cantidad": unidades[n],
            "importeNeto": precio[n],
            "descripcion": producto[n]
        }
        lineasNorm.append(json2)
        altaDocs(letra[n], puestoVenta[n], facturas[n], fecha_convertida, caea[n], descripcion_medicamento, lineasNorm, id_afip, idalicuota, iva[n])
        recorrer_carpetas(facturas[n], id_afip, puestoVenta[n], letra[n])
        print("")
    time.sleep(3)
        
    lineasNorm.clear()
    n += 1

    
    
        
        
    



