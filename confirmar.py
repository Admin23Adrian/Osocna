from autenticarse import loguearse
import time
import requests
import openpyxl
from openpyxl import load_workbook

excel = openpyxl.load_workbook("C:/Users/aalarcon/Desktop/osocna.xlsx")
hoja = excel["osocna"]

token = loguearse()

def confirmar(id_afip, punto_venta, factura, letra):
    letra_nueva = letra.strip()
    url_confirmar = "https://extranet.osocna.com.ar/Backend/api/ComprobantesExtranet/tipoComprobante/"+ str(id_afip) +"/puntoVenta/"+ str(punto_venta) +"/nroComprobante/"+ str(factura) +"/tipoEmision/"+ str(letra_nueva) +"/Confirmar"
    headers_confirmar = {"Authorization":token}
    payload = {}
    request = requests.post(url = url_confirmar, headers = headers_confirmar, data = payload)

    response = request.text
    print(response, request.status_code)

cantidad = len(hoja["A"])

idAfip = None
for i in range(3, cantidad+1):
    factura = hoja.cell(row=i, column=5).value
    puesto_v = hoja.cell(row=i, column=3).value
    letra = hoja.cell(row=i, column=4).value
    
    if puesto_v == 44:
        idAfip = 1
    elif puesto_v == 50:
        idAfip = 3
    
    confirmar(idAfip, puesto_v, factura, letra)
    time.sleep(2)
    print("")

